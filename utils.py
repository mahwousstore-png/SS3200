"""
utils.py
Helper functions for reading and writing Excel files (Salla template format).
"""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_excel(file_obj) -> pd.DataFrame:
    """
    Parse the Salla-format Excel file.
    Handles the two-row header (row 1 = 'بيانات المنتج', row 2 = actual headers).
    Returns a cleaned DataFrame.
    """
    # Read with the second row as header
    df = pd.read_excel(file_obj, header=1)
    
    # Drop completely empty rows
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    
    return df


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Export DataFrame back to Excel preserving the Salla template structure.
    Returns bytes buffer.
    """
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Write with header starting at row 2 (leave row 1 for title)
        df.to_excel(writer, index=False, sheet_name="Salla Products Template Sheet", startrow=1)
        
        wb = writer.book
        ws = writer.sheets["Salla Products Template Sheet"]
        
        # Add title row (row 1)
        ws.insert_rows(1)
        ws["A1"] = "بيانات المنتج"
        ws["A1"].font = Font(bold=True, size=13, name="Tajawal")
        ws["A1"].fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
        ws.row_dimensions[1].height = 30
        
        # Style header row (row 2)
        for cell in ws[2]:
            cell.font = Font(bold=True, size=10, name="Tajawal")
            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40
        
        # Auto-adjust column widths (rough)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value or "")
                    # Don't stretch by HTML content
                    display_len = min(len(val), 50)
                    if display_len > max_len:
                        max_len = display_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 55))
        
        # RTL sheet direction
        ws.sheet_view.rightToLeft = True
    
    buffer.seek(0)
    return buffer.getvalue()


def extract_product_info_from_row(row: pd.Series) -> dict:
    """Extract standardized product info from a DataFrame row."""
    
    def safe_get(key):
        val = row.get(key)
        if val is None or (isinstance(val, float) and str(val) == "nan"):
            return ""
        return str(val).strip()
    
    return {
        "name": safe_get("أسم المنتج"),
        "brand": safe_get("الماركة"),
        "category": safe_get("تصنيف المنتج"),
        "sku": safe_get("رمز المنتج sku"),
        "price": safe_get("سعر المنتج"),
        "status": safe_get("حالة المنتج"),
        "weight": safe_get("الوزن"),
    }
